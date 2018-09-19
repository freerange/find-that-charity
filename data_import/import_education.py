import argparse
import os
import csv
import datetime
import re

from elasticsearch import Elasticsearch
from openpyxl import load_workbook

from import_charities import save_to_elasticsearch

AREA_TYPES = {
    "E00": "OA",
    "E01": "LSOA",
    "E02": "MSOA",
    "E04": "PAR",
    "E05": "WD",
    "E06": "UA",
    "E07": "NMD",
    "E08": "MD",
    "E09": "LONB",
    "E10": "CTY",
    "E12": "RGN/GOR",
    "E14": "WPC",
    "E15": "EER",
    "E21": "CANNET",
    "E22": "CSP",
    "E23": "PFA",
    "E25": "PUA",
    "E26": "NPARK",
    "E28": "REGD",
    "E29": "REGSD",
    "E30": "TTWA",
    "E31": "FRA",
    "E32": "LAC",
    "E33": "WZ",
    "E36": "CMWD",
    "E37": "LEP",
    "E38": "CCG",
    "E39": "NHSAT",
    "E41": "CMLAD",
    "E42": "CMCTY",
    "N00": "SA",
    "N06": "WPC",
    "S00": "OA",
    "S01": "DZ",
    "S02": "IG",
    "S03": "CHP",
    "S05": "ROA - CPP",
    "S06": "ROA - Local",
    "S08": "HB",
    "S12": "CA",
    "S13": "WD",
    "S14": "WPC",
    "S16": "SPC",
    "S22": "TTWA",
    "W00": "OA",
    "W01": "LSOA",
    "W02": "MSOA",
    "W03": "USOA",
    "W04": "COM",
    "W05": "WD",
    "W06": "UA",
    "W07": "WPC",
    "W09": "NAWC",
    "W14": "CDRP",
    "W20": "REGD",
    "W21": "REGSD",
    "W22": "TTWA",
    "W30": "AgricSmall",
    "W33": "CFA",
    "W35": "WZ",
    "W39": "CMWD",
    "W40": "CMLAD",
    "W41": "CMCTY",
}

REGION_CONVERT = {
    "A": "E12000001",
    "B": "E12000002",
    "D": "E12000003",
    "E": "E12000004",
    "F": "E12000005",
    "G": "E12000006",
    "H": "E12000007",
    "J": "E12000008",
    "K": "E12000009",
}

def import_gias(orgs,
                datafile=os.path.join("data", "gias_england.csv"),
                es_index="charitysearch",
                es_type="organisation",
                debug=False):
    with open(datafile) as a:
        csvreader = csv.DictReader(a)
        rcount = 0
        for row in csvreader:
            row = clean_gias(row)
            row["_index"] = es_index
            row["_type"] = es_type
            row["_op_type"] = "index"
            orgs[row["_id"]] = row
            rcount += 1
            if rcount % 10000 == 0:
                print('\r', "[GIAS] {} organisations added or updated from {}".format(rcount, datafile), end='')
            if debug and rcount > 500:
                break
    print('\r', "[GIAS] {} organisations added or updated from {}".format(rcount, datafile))

    return orgs

def import_scot(orgs,
                datafile=os.path.join("data", "schools_scotland.xlsx"),
                es_index="charitysearch",
                es_type="organisation",
                debug=False,
                skip_rows = 5):
    # @TODO: Set all to inactive before importing new data
    wb = load_workbook(datafile, read_only=True)
    rcount = 0
    latest_sheet = wb[sorted(
        [s for s in wb.sheetnames if s.startswith("Open at")])[-1]]
    headers = {}
    seen_blank_row = False
    for k, row in enumerate(latest_sheet.rows):
        if k < skip_rows or seen_blank_row:
            continue
        elif k == skip_rows:
            previous_overtitle = None
            header_names = []
            for c in row:
                if c.value:
                    title = str(c.value)

                    # get the row before to find the heading for this title
                    overtitle = latest_sheet.cell(k, c.column).value
                    if overtitle is None:
                        overtitle = previous_overtitle
                    else:
                        previous_overtitle = overtitle

                    # we actually only need three of them
                    if overtitle:
                        if overtitle.startswith("Pupil rolls"):
                            overtitle = "Pupil rolls"
                        elif overtitle.startswith("Teachers"):
                            overtitle = "Teachers FTE"
                        elif overtitle.startswith("School type"):
                            overtitle = "School type"
                        else:
                            overtitle = None

                    header_names.append(slugify("{} {}".format(
                        overtitle if overtitle else "", title)))

            headers = dict(zip(
                [c.column for c in row if c.value],  # header column numbers
                header_names
            ))
            continue
        elif row[0].value is None:
            seen_blank_row = True
            continue

        row_data = {}
        for i, c in enumerate(row):
            if i+1 in headers:
                v = c.value
                if v in ["", ".", "N/A", "0", 0]:
                    v = None
                row_data[headers[i+1]] = v
        row_data = clean_scot(row_data)
        row_data["_index"] = es_index
        row_data["_type"] = es_type
        row_data["_op_type"] = "index"
        orgs[row_data["_id"]] = row_data
        rcount += 1
        if rcount % 10000 == 0:
            print('\r', "[SCOTLAND] {} organisations added or updated from {}".format(rcount, datafile), end='')
        if debug and rcount > 500:
            break
    print('\r', "[SCOTLAND] {} organisations added or updated from {}".format(rcount, datafile))

    return orgs

def import_ni(orgs,
              datafile=os.path.join("data", "schools_ni.csv"),
              es_index="charitysearch",
              es_type="organisation",
              debug=False):
    with open(datafile) as a:
        csvreader = csv.DictReader(a)
        rcount = 0
        for row in csvreader:
            row = clean_ni(row)
            if row["_id"] == "GB-NIEDU-UNKNOWN":
                continue
            row["_index"] = es_index
            row["_type"] = es_type
            row["_op_type"] = "index"
            orgs[row["_id"]] = row
            rcount += 1
            if rcount % 10000 == 0:
                print('\r', "[NI] {} organisations added or updated from {}".format(rcount, datafile), end='')
            if debug and rcount > 500:
                break
    print('\r', "[NI] {} organisations added or updated from {}".format(rcount, datafile))

    return orgs


def slugify(value):
    value = value.lower()
    value = re.sub(r'\([0-9]+\)', "_", value).strip("_") # replace values in brackets
    value = re.sub(r'[^0-9A-Za-z]+', "_", value).strip("_") # replace any non-alphanumeric characters
    return value

def clean_gias(record):

    # clean blank values
    for f in record.keys():
        if record[f] == "":
            record[f] = None

    # dates
    date_fields = ["OpenDate", "CloseDate"]
    for f in date_fields:
        try:
            if record.get(f):
                record[f] = datetime.datetime.strptime(record.get(f), "%d-%m-%Y")
        except ValueError:
            record[f] = None

    # org ids:
    org_ids = ["GB-EDU-{}".format(record.get("URN"))]
    if record.get("UKPRN"):
        org_ids.append("GB-UKPRN-{}".format(record.get("UKPRN")))
    if record.get("EstablishmentNumber") and record.get("LA (code)"):
        org_ids.append("GB-LAESTAB-{}/{}".format(
            record.get("LA (code)").rjust(3, "0"),
            record.get("EstablishmentNumber").rjust(4, "0"),
        ))

    return {
        "_id": "GB-EDU-{}".format(record.get("URN")),
        "name": record.get("EstablishmentName"),
        "charityNumber": None,
        "companyNumber": None,
        "streetAddress": record.get("Street"),
        "addressLocality": record.get("Locality"),
        "addressRegion": record.get("Address3"),
        "addressCountry": record.get("Country (name)"),
        "postalCode": record.get("Postcode"),
        "telephone": record.get("TelephoneNum"),
        "alternateName": [],
        "email": None,
        "description": None,
        "organisationType": [
            "Education",
            record.get("EstablishmentTypeGroup (name)"),
            record.get("TypeOfEstablishment (name)"),
        ],
        "url": record.get("SchoolWebsite"),
        "location": get_locations(record),
        "latestIncome": None,
        "dateModified": datetime.datetime.now(),
        "dateRegistered": record.get("OpenDate"),
        "dateRemoved": record.get("CloseDate"),
        "active": record.get("EstablishmentStatus (name)") != "Closed",
        "parent": record.get("PropsName"),
        "orgIDs": org_ids,
    }

def clean_scot(record):
    scot_las = {
        "Aberdeen City": "S12000033",
        "Aberdeenshire": "S12000034",
        "Angus": "S12000041",
        "Argyll & Bute": "S12000035",
        "Clackmannanshire": "S12000005",
        "Dumfries & Galloway": "S12000006",
        "Dundee City": "S12000042",
        "East Ayrshire": "S12000008",
        "East Dunbartonshire": "S12000045",
        "East Lothian": "S12000010",
        "East Renfrewshire": "S12000011",
        "Edinburgh City": "S12000036",
        "Falkirk": "S12000014",
        "Fife": "S12000015",
        "Glasgow City": "S12000046",
        "Highland": "S12000017",
        "Inverclyde": "S12000018",
        "Midlothian": "S12000019",
        "Moray": "S12000020",
        "Na h-Eileanan Siar": "S12000013",
        "North Ayrshire": "S12000021",
        "North Lanarkshire": "S12000044",
        "Orkney Islands": "S12000023",
        "Perth & Kinross": "S12000024",
        "Renfrewshire": "S12000038",
        "Scottish Borders": "S12000026",
        "Shetland Islands": "S12000027",
        "South Ayrshire": "S12000028",
        "South Lanarkshire": "S12000029",
        "Stirling": "S12000030",
        "West Dunbartonshire": "S12000039",
        "West Lothian": "S12000040",
    }

    org_types = [
        "Education",
        record.get("centre_type") + " School",
    ]
    for f in ["school_type_primary", "school_type_secondary", "school_type_special"]:
        if record.get(f):
            org_types.append(record[f] + " School")
    org_id = "GB-SCOTEDU-{}".format(record.get("seedcode"))
    locations = []
    if scot_las.get(record.get("la_name")):
        code = scot_las.get(record.get("la_name"))
        locations.append({
            "id": code,
            "name": record.get("la_name"),
            "geoCode": code,
            "geoCodeType": AREA_TYPES.get(code[0:3], code),
        })
    
    return {
        "_id": org_id,
        "name": record.get("school_name"),
        "charityNumber": None,
        "companyNumber": None,
        "streetAddress": record.get("address_1"),
        "addressLocality": record.get("address_2"),
        "addressRegion": record.get("address_3"),
        "addressCountry": "Scotland",
        "postalCode": record.get("post_code"),
        "telephone": record.get("phone"),
        "alternateName": [],
        "email": record.get("e_mail"),
        "description": None,
        "organisationType": org_types,
        "url": None,
        "location": locations,
        "latestIncome": None,
        "dateModified": datetime.datetime.now(),
        "dateRegistered": None,
        "dateRemoved": None,
        "active": True,
        "parent": None,
        "orgIDs": [
            org_id
        ],
    }


def clean_ni(record):

    # clean blank values
    for f in record.keys():
        record[f] = record[f].strip()
        if record[f] == "":
            record[f] = None

    # dates
    date_fields = ["Date Closed"]
    for f in date_fields:
        try:
            if record.get(f):
                record[f] = datetime.datetime.strptime(
                    record.get(f), "%d-%m-%Y")
        except ValueError:
            record[f] = None

    # org ids:
    org_id = "GB-NIEDU-{}".format(record.get("Institution Reference Number"))

    address = ", ".join([
        record.get("Address Line {}".format(f))
        for f in [1,2,3] 
        if record.get("Address Line {}".format(f))
    ])

    return {
        "_id": org_id,
        "name": record.get("Institution Name").strip(),
        "charityNumber": None,
        "companyNumber": None,
        "streetAddress": address,
        "addressLocality": record.get("Town"),
        "addressRegion": record.get("Count"),
        "addressCountry": "Northern Ireland",
        "postalCode": record.get("Postcode"),
        "telephone": record.get("Telephone"),
        "alternateName": [],
        "email": record.get("Email"),
        "description": None,
        "organisationType": [
            "Education",
            record.get("Management"),
            record.get("Type", "") + " School",
        ],
        "url": None,
        "location": [],
        "latestIncome": None,
        "dateModified": datetime.datetime.now(),
        "dateRegistered": None,
        "dateRemoved": record.get("Date Closed"),
        "active": record.get("Status") == "Open",
        "parent": None,
        "orgIDs": [org_id],
    }

def get_locations(record):
    location_fields = ["GOR", "DistrictAdministrative", "AdministrativeWard",
                       "ParliamentaryConstituency", "UrbanRural", "MSOA", "LSOA"]
    locations = []
    for f in location_fields:
        code = record.get(f+" (code)", "")
        name = record.get(f+" (name)", "")

        if name == "" and code == "":
            continue

        if f == "GOR":
            code = REGION_CONVERT.get(code, code)

        if code == "":
            code = name

        locations.append({
            "id": code,
            "name": record.get(f+" (name)"),
            "geoCode": code,
            "geoCodeType": AREA_TYPES.get(code[0:3], f),
        })

    return locations


def main():

    parser = argparse.ArgumentParser(description='Import education data into elasticsearch')

    parser.add_argument('--folder', type=str, default='data',
                        help='Root path of the data folder.')

    # elasticsearch options
    parser.add_argument('--es-host', default="localhost", help='host for the elasticsearch instance')
    parser.add_argument('--es-port', default=9200, help='port for the elasticsearch instance')
    parser.add_argument('--es-url-prefix', default='', help='Elasticsearch url prefix')
    parser.add_argument('--es-use-ssl', action='store_true', help='Use ssl to connect to elasticsearch')
    parser.add_argument('--es-index', default='charitysearch', help='index used to store charity data')
    parser.add_argument('--es-type', default='organisation', help='type used to store organisation data')

    parser.add_argument('--skip-gias', action='store_true',
                        help='Don\'t fetch data from English schools list.')
    parser.add_argument('--skip-scot', action='store_true',
                        help='Don\'t fetch data from Scottish schools list.')
    parser.add_argument('--skip-ni', action='store_true',
                        help='Don\'t fetch data from NI schools list.')

    parser.add_argument('--debug', action='store_true', help='')

    args = parser.parse_args()

    es = Elasticsearch(host=args.es_host, port=args.es_port, url_prefix=args.es_url_prefix, use_ssl=args.es_use_ssl)

    potential_env_vars = [
        "ELASTICSEARCH_URL",
        "ES_URL",
        "BONSAI_URL"
    ]
    for e_v in potential_env_vars:
        if os.environ.get(e_v):
            es = Elasticsearch(os.environ.get(e_v))
            break

    if not es.ping():
        raise ValueError("Elasticsearch connection failed")

    data_files = {
        "gias": os.path.join(args.folder, "gias_england.csv"),
        "scot": os.path.join(args.folder, "schools_scotland.xlsx"),
        "ni": os.path.join(args.folder, "schools_ni.csv"),
    }

    orgs = {}
    if not args.skip_gias:
        orgs = import_gias(orgs, datafile=data_files["gias"], es_index=args.es_index, es_type=args.es_type, debug=args.debug)

    if not args.skip_scot:
        orgs = import_scot(orgs, datafile=data_files["scot"], es_index=args.es_index, es_type=args.es_type, debug=args.debug)

    if not args.skip_ni:
        orgs = import_ni(orgs, datafile=data_files["ni"], es_index=args.es_index, es_type=args.es_type, debug=args.debug)

    if args.debug:
        import random
        random_keys = random.choices(list(orgs.keys()), k=10)
        for r in random_keys:
            print(r, orgs[r])

    save_to_elasticsearch(orgs, es, args.es_index)

if __name__ == '__main__':
    main()
